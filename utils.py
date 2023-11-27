import os
import openpyxl

from slugify.slugify import slugify
from settings import settings
from models import Book


# Save file to uploads folder
async def save_file_to_uploads(file, filename, file_id):
    file_path = get_path_to_file(filename, file_id)
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    with open(file_path, "wb") as uploaded_file:
        file_content = await file.read()
        uploaded_file.write(file_content)


# Delete file from uploads folder
def delete_file_from_uploads(file_name):
    try:
        os.remove(file_name)
    except FileNotFoundError as e:
        print(e)


# Format filename
def format_filename(title: str):
    return slugify(title)


def get_path_to_file(filename: str, file_id):
    return os.path.join(settings.UPLOADED_FILES_PATH, str(file_id), f"{filename}.pdf")


def get_authors_list_from_xls(file_path):
    workbook = openpyxl.load_workbook(file_path, read_only=True)
    first_sheet = workbook.worksheets[0]
    # workbook.close()

    return [cell for row in first_sheet.iter_rows(values_only=True) for cell in row]  # type: ignore


def get_titles_list_from_xls(file_path):
    workbook = openpyxl.load_workbook(file_path, read_only=True)
    second_sheet = workbook.worksheets[1]
    # workbook.close()

    return [
        cell for row in second_sheet.iter_rows(values_only=True) for cell in row
    ]  # type: ignore


def create_img_tag(filename: str):
    return (f'''
        <div style='margin: 10px;'>
            <img src='/{filename}' alt='Page' width='700' height='900'>
        </div>
    ''')


def create_html_page(book: Book) -> str:
    images_folder = os.path.join(settings.UPLOADED_FILES_PATH, str(book.id), "img")

    image_paths = [
        os.path.join(images_folder, filename)
        for filename in os.listdir(images_folder)
        if filename.endswith(".jpg")
    ]

    image_tags = [create_img_tag(path) for path in image_paths]

    return f"""
        <!DOCTYPE html>
            <html lang="en">
            <head>
                <meta charset="UTF-8">
                <meta name="viewport" content="width=device-width, initial-scale=1.0">
                <title>{book.title}</title>
            </head>
            <body style="display: flex; flex-direction: column; align-items: center; justify-content: center; margin: 0;">

            {' '.join(image_tags)}

            </body>
            </html>
        """
